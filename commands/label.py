import Command_Engine
import os
import glob
import traceback
import numpy as np
import matplotlib
matplotlib.use('Agg')

import matplotlib.cm as cm
import matplotlib.colors as mcolors
from Bio import AlignIO
from Bio.Align import MultipleSeqAlignment 
import SSN_Config as cfg
import SSN_Utils as utils

def print_help():
    print("""
    Differential Labeling & Statistics Tool
    =======================================
    Generates a comprehensive XLSX report comparing the sequence properties and 
    conserved residues of each subset against the global dataset. Output is saved 
    to the 'Results/Cluster_Label/' directory.

    * PREREQUISITES: 
      1. A Multiple Sequence Alignment (MSA) must be loaded.
      2. A Reference Sequence must be set (use the 'reference' command).

    Usage: label [TARGET] [GLOBAL_MAX] [CLUSTER_MIN] [GLOBAL_MIN]
       or: label [TARGET] [key value] [<key 2> <value 2> ...]

    Targets (Default: clusters):
      clusters : Analyzes all defined topology clusters AND any custom groups.
      groups   : Analyzes ONLY custom groups (topology clusters not required).

    Arguments (Accepts decimals '0.3' or percentages '30%'):
      gmax (Global Max)   : Default 30%. Max frequency a residue can have in the 
                            GLOBAL alignment to be considered "Subset Specific".
      cmin (Cluster Min)  : Default 95%. Min frequency a residue must have WITHIN 
                            a subset to be reported as conserved.
      gmin (Global Min)   : Default 95%. Min frequency a residue must have in the 
                            GLOBAL dataset to be reported as globally conserved.

    Examples:
      label                       (Runs using default parameters)
      label 0.4 0.9 0.95          (Positional: gmax=40%, cmin=90%, gmin=95%)
      label groups cmin 90%       (Keyword: Analyzes groups, sets cmin to 90%)
      
    Note: Do not mix positional numbers after using keywords.
    """)

def parse_percentage(val_str):
    try:
        clean_str = val_str.replace('%', '')
        val = float(clean_str)
        if val > 1.0: return val / 100.0
        return val
    except ValueError: return None

def get_sequence_stats(aln):
    lengths = []
    gap_chars = set(cfg.GAP_CHARS)
    for record in aln:
        seq_str = str(record.seq)
        ungapped_len = sum(1 for c in seq_str if c not in gap_chars)
        lengths.append(ungapped_len)
    if not lengths: return 0, 0, 0.0, 0.0
    arr = np.array(lengths)
    return int(np.min(arr)), int(np.max(arr)), np.mean(arr), np.std(arr)

def run(viewer, args):
    if args and args[0].lower() == 'reset':
        Command_Engine.execute_reset(viewer, ["clusters"])
        return

    try:
        if viewer.alignment.aln is None:
            viewer.console_text.text = "Error: Global Alignment not loaded."
            print("Error: Global Alignment not loaded.")
            return
            
        if not getattr(viewer, 'active_reference', None):
            viewer.console_text.text = "Error: No Reference Set. Use 'reference <ID>' first."
            print("Error: No Reference Set. Use 'reference <ID>' first.")
            return

        if args and args[0].lower() in ['help', '-h', '-?']:
            print_help()
            if hasattr(viewer, 'console_text'):
                viewer.console_text.text = "Help information printed to the terminal"
            return

        # --- Parameters ---
        global_max = 0.30
        cluster_min = 0.95
        global_min = 0.95 
        forced_target = "clusters"

        valid_keys = {"gmax", "global_max", "g_max", "cmin", "cluster_min", "c_min", "gmin", "global_min", "g_min"}
        valid_targets = {"cluster", "clusters", "group", "groups"}
        
        positional_args = []
        keyword_args = {}
        keyword_mode = False
        
        i = 0
        while i < len(args):
            arg = args[i].lower()
                
            # Catch targets anywhere in the command
            if arg in valid_targets:
                forced_target = "clusters" if arg in ["cluster", "clusters"] else "groups"
                i += 1
                continue
                
            # Catch space-separated keywords
            if arg in valid_keys:
                keyword_mode = True
                if i + 1 >= len(args):
                    msg = f"Error: Missing numerical value for '{arg}'."
                    viewer.console_text.text = msg
                    print(msg)
                    return
                
                val_str = args[i+1]
                val = parse_percentage(val_str)
                if val is None:
                    msg = f"Error: Invalid percentage '{val_str}' for '{arg}'."
                    viewer.console_text.text = msg
                    print(msg)
                    return
                
                # Standardize key names
                if arg in ["gmax", "global_max", "g_max"]: key_name = "gmax"
                elif arg in ["cmin", "cluster_min", "c_min"]: key_name = "cmin"
                elif arg in ["gmin", "global_min", "g_min"]: key_name = "gmin"
                
                if key_name in keyword_args:
                    msg = f"Error: Duplicate assignment for '{key_name}'."
                    viewer.console_text.text = msg
                    print(msg)
                    return
                    
                keyword_args[key_name] = val
                i += 2
                continue
                
            # If it is not a target or a keyword, it MUST be a positional number
            if keyword_mode:
                msg = f"Error: Ambiguous input. Positional argument '{arg}' found after keywords."
                viewer.console_text.text = msg
                print(msg)
                return
                
            val = parse_percentage(arg)
            if val is not None:
                positional_args.append(val)
            else:
                msg = f"Error: Unrecognized argument or invalid number '{arg}'."
                viewer.console_text.text = msg
                print(msg)
                return
                
            i += 1

        # Map positionals strictly to order: 1. gmax, 2. cmin, 3. gmin
        pos_map = ["gmax", "cmin", "gmin"]
        if len(positional_args) > 3:
            msg = "Error: Too many positional numerical arguments."
            viewer.console_text.text = msg
            print(msg)
            return
            
        for idx, p_val in enumerate(positional_args):
            target_key = pos_map[idx]
            if target_key in keyword_args:
                msg = f"Error: Ambiguous input. '{target_key}' defined both positionally and via keyword."
                viewer.console_text.text = msg
                print(msg)
                return
            keyword_args[target_key] = p_val
            
        # Apply final parsed variables (falling back to defaults)
        global_max = keyword_args.get("gmax", 0.30)
        cluster_min = keyword_args.get("cmin", 0.95)
        global_min = keyword_args.get("gmin", 0.95)

        # --- Validations ---
        if forced_target == "clusters" and viewer.cluster_labels is None:
            viewer.console_text.text = "Error: Run 'cluster' first."
            print("Error: Run 'cluster' first to use cluster mode.")
            return
            
        if forced_target == "groups" and getattr(viewer, 'group_labels', None) is None:
            viewer.console_text.text = "Error: No groups defined."
            print("Error: No groups defined. Use the 'group' command first.")
            return

        # --- 1. Global Statistics ---
        print("Calculating Global Stats...")
        g_stats = viewer.alignment.calculate_frequencies(viewer.alignment.col_to_label)
        total_global_seqs = len(viewer.alignment.aln)
        g_min, g_max, g_avg, g_std = get_sequence_stats(viewer.alignment.aln)

        # --- 2. Resolve Base Directories ---
        import re
        hdf5_base = os.path.basename(getattr(cfg, 'INPUT_HDF5', ''))
        fasta_file = getattr(cfg, 'NODE_FASTA_FILE', None)
        fasta_base = os.path.splitext(os.path.basename(fasta_file))[0] if fasta_file else getattr(cfg, 'SEQUENCE_SET', 'Network')
        
        match = re.search(r'(\[.*?\])', hdf5_base)
        if match:
            model_str = f"_{match.group(1)}"
        else:
            hdf5_no_ext = hdf5_base[:-3] if hdf5_base.endswith(".h5") else os.path.splitext(hdf5_base)[0]
            stripped = re.sub(r'_(network|evalue)$', '', hdf5_no_ext, flags=re.IGNORECASE)
            old_match = re.search(r'_(e[0-9]+_.*|blast.*)$', stripped, flags=re.IGNORECASE)
            model_str = f"_{old_match.group(1)}" if old_match else ""
            
        lvl1_name = f"{fasta_base}{model_str}"
        is_blast = "EValue" in hdf5_base or "Evalue" in hdf5_base or "blast" in hdf5_base.lower()
        if not is_blast:
            norm_m = getattr(cfg, 'NORM_MODE', None)
            if norm_m: lvl1_name += f"_{norm_m}"
            score_m = getattr(cfg, 'ALIGNMENT_SCORE', None)
            if score_m: lvl1_name += f"_{score_m}"
            
        lvl2_name_base = ""
        top_val = getattr(cfg, 'TOP_EDGE_PERCENT', None)
        if top_val is not None and str(top_val).strip() != "None":
            try: lvl2_name_base += f"Top{float(top_val)}Pct"
            except: pass
        else:
            thresh = getattr(cfg, 'SIMILARITY_THRESHOLD', 0.0)
            try: lvl2_name_base += f"Score{float(thresh)}"
            except: pass
            
        if forced_target == "clusters":
            if getattr(viewer, 'last_cluster_params', None):
                c_mode_param, c_min_param = viewer.last_cluster_params
                if lvl2_name_base:
                    lvl2_name = f"{lvl2_name_base}_{c_mode_param}_Min{c_min_param}"
                else:
                    lvl2_name = f"{c_mode_param}_Min{c_min_param}"
            else:
                lvl2_name = lvl2_name_base
        else:
            lvl2_name = lvl2_name_base

        # --- 3. Prepare Tasks ---
        tasks = [] 
        
        print(f"Splitting Global Alignment for {forced_target.upper()}...")
        
        if forced_target == "clusters":
            aln_idx_to_cid = {}
            for i, simple_h in enumerate(viewer.full_headers):
                if i >= len(viewer.cluster_labels): break
                cid = viewer.cluster_labels[i]
                if viewer.alignment.seq_map and simple_h in viewer.alignment.seq_map:
                    aln_idx = viewer.alignment.seq_map[simple_h]
                    aln_idx_to_cid[aln_idx] = cid

            clusters_records = {}
            for i, record in enumerate(viewer.alignment.aln):
                if i in aln_idx_to_cid:
                    found_cid = aln_idx_to_cid[i]
                    if found_cid != -1:
                        if found_cid not in clusters_records: clusters_records[found_cid] = []
                        clusters_records[found_cid].append(record)
            
            for cid in sorted(clusters_records.keys()):
                sub_aln = MultipleSeqAlignment(clusters_records[cid])
                tasks.append(('cluster', cid, sub_aln, viewer.alignment.col_to_label))
        
        # Group splitting
        if getattr(viewer, 'group_labels', None):
            aln_idx_to_groups = {}
            for i, simple_h in enumerate(viewer.full_headers):
                if i >= len(viewer.group_labels): break
                groups = viewer.group_labels[i]
                if groups and viewer.alignment.seq_map and simple_h in viewer.alignment.seq_map:
                    aln_idx = viewer.alignment.seq_map[simple_h]
                    aln_idx_to_groups[aln_idx] = groups
            
            groups_records = {}
            for i, record in enumerate(viewer.alignment.aln):
                if i in aln_idx_to_groups:
                    for g_name in aln_idx_to_groups[i]:
                        if g_name not in groups_records: groups_records[g_name] = []
                        groups_records[g_name].append(record)
            
            for g_name in sorted(groups_records.keys()):
                sub_aln = MultipleSeqAlignment(groups_records[g_name])
                tasks.append(('group', g_name, sub_aln, viewer.alignment.col_to_label))

        # --- 4. Process Tasks ---
        master_labels = set()
        cluster_results = []
        
        try:
            cmap = cm.get_cmap('tab20')
        except AttributeError:
            cmap = matplotlib.colormaps['tab20']

        for entity_type, entity_id, c_aln, c_map in tasks:
            try:
                # Calculate sequence stats
                c_size = len(c_aln)
                c_min_len, c_max_len, c_avg_len, c_std_len = get_sequence_stats(c_aln)

                # Calculate Residue Frequencies (using global map implicitly)
                c_stats = viewer.alignment.calculate_frequencies(c_map, exclude=[], aln=c_aln)
                c_dict = {}
                c_occ_dict = {} 
                
                for lbl in utils.sort_labels(c_stats.keys()):
                    if lbl not in c_stats: continue
                    c_aa, c_freq, c_occ = c_stats[lbl]
                    
                    c_occ_dict[lbl] = c_occ 
                    if c_freq < cluster_min: continue
                    
                    is_interesting = False
                    if lbl in g_stats:
                        g_aa, g_freq, _ = g_stats[lbl]
                        if g_aa == c_aa:
                            if g_freq < global_max: is_interesting = True
                        else: is_interesting = True 
                    
                    if is_interesting:
                        c_dict[lbl] = {"text": f"{c_aa}{lbl}", "occ": c_occ}
                        master_labels.add(lbl)
                
                # Format Output Styling
                if entity_type == 'cluster':
                    hex_code = mcolors.to_hex(cmap(entity_id % 20))
                    name_str = f"Cluster {entity_id}"
                    sort_key = (0, entity_id)
                else:
                    hex_code = "-"
                    name_str = f"Group {entity_id}"
                    # Sort groups by count descending (-c_size), then alphabetically
                    sort_key = (1, -c_size, str(entity_id))

                cluster_results.append({
                    "type": entity_type,
                    "id": entity_id,
                    "name": name_str,
                    "sort_key": sort_key,
                    "count": c_size,
                    "hex": hex_code,
                    "min": c_min_len,
                    "max": c_max_len,
                    "avg": c_avg_len,
                    "std": c_std_len,
                    "data": c_dict,
                    "occ_data": c_occ_dict
                })
                    
            except Exception as e: 
                print(f"Skipping {entity_type} {entity_id} due to error: {e}")
                continue

        # --- 5. Export XLSX ---
        out_dir = cfg.CLUSTER_LABEL_DIR
        if not os.path.exists(out_dir): os.makedirs(out_dir)
        
        # Build clean filename
        if lvl2_name:
            out_filename = f"{lvl1_name}_{lvl2_name}_Gmax{int(global_max*100)}_Cmin{int(cluster_min*100)}_Gmin{int(global_min*100)}.xlsx"
        else:
            out_filename = f"{lvl1_name}_Gmax{int(global_max*100)}_Cmin{int(cluster_min*100)}_Gmin{int(global_min*100)}.xlsx"
            
        out_path = os.path.join(out_dir, out_filename)

        global_list = []
        for lbl in utils.sort_labels(g_stats.keys()):
            aa, freq, occ = g_stats[lbl]
            if freq > global_min:
                global_list.append(f"{aa}{lbl}")

        sorted_cols = utils.sort_labels(list(master_labels))
        cluster_results.sort(key=lambda x: x["sort_key"])
        all_occ_labels = utils.sort_labels(list(g_stats.keys()))
        
        ref_display = getattr(viewer, 'resolved_ref_full', None) or viewer.active_reference or "None"
        
        try:
            import openpyxl
            from openpyxl.styles import PatternFill
        except ImportError:
            viewer.console_text.text = "Error: 'openpyxl' is required for XLSX export. Run: pip install openpyxl"
            print("Error: openpyxl not installed.")
            return

        try:
            wb = openpyxl.Workbook()
            
            # ==========================================
            # TAB 1: Subset Specific Matrix
            # ==========================================
            ws1 = wb.active
            ws1.title = "Subset Stats"
            
            # Write Metadata 
            ws1.append([f"Filename: {out_filename}"])
            ws1.append([f"Reference: {ref_display}"])
            ws1.append([f"Global Conserved (>{int(global_min*100)}%)"])
            ws1.append(global_list if global_list else ["None"])
            ws1.append([])
            
            # Write Headers 
            ws1.append(["Subset Specific Matrix"])
            headers1 = ["Subset Name", "Count", "Hex Color", "Min Len", "Max Len", "Avg Len", "Std Dev", ""] + [f"#{c}" for c in sorted_cols]
            ws1.append(headers1)
            
            try:
                occ_cmap1 = cm.get_cmap('Reds_r')
            except AttributeError:
                occ_cmap1 = matplotlib.colormaps['Reds_r']
                
            def get_occ_fill1(occ_value):
                rgba = occ_cmap1(occ_value)
                hex_color = mcolors.to_hex(rgba)[1:].upper()
                return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

            # Write Global Row 
            global_freq_row1 = [
                "Global Stats", 
                total_global_seqs, 
                "-",
                g_min, 
                g_max, 
                round(g_avg, 1), 
                round(g_std, 1),
                ""
            ]
            
            g_occ_dict1 = {}
            for col in sorted_cols:
                if col in g_stats:
                    g_aa, g_freq, g_occ = g_stats[col]
                    global_freq_row1.append(f"{g_aa} ({int(g_freq*100)}%)")
                    g_occ_dict1[col] = g_occ
                else: 
                    global_freq_row1.append("-")
                    
            ws1.append(global_freq_row1)
            g_row_idx1 = ws1.max_row
            
            for c_idx, col in enumerate(sorted_cols):
                if col in g_occ_dict1:
                    col_letter_idx = c_idx + 9 
                    ws1.cell(row=g_row_idx1, column=col_letter_idx).fill = get_occ_fill1(g_occ_dict1[col])
                    
            ws1.append([]) # Blank row below Global Stats

            # Write Subset Rows
            last_type = None
            for res in cluster_results:
                if last_type == 'cluster' and res['type'] == 'group':
                    ws1.append([]) # Blank row separating Clusters and Groups
                last_type = res['type']
                row1 = [
                    res['name'], 
                    res['count'], 
                    res['hex'],
                    res['min'],
                    res['max'],
                    round(res['avg'], 1),
                    round(res['std'], 1),
                    ""
                ]
                
                row_occs1 = {}
                for c_idx, col in enumerate(sorted_cols): 
                    if col in res['data']:
                        row1.append(res['data'][col]["text"])
                    else:
                        row1.append("")
                        
                    if col in res['occ_data']:
                        row_occs1[c_idx + 9] = res['occ_data'][col]
                    else:
                        row_occs1[c_idx + 9] = 0.0 
                        
                ws1.append(row1)
                current_row1 = ws1.max_row
                
                if res['hex'] != "-":
                    hex_val = res['hex'].replace("#", "").upper()
                    ws1.cell(row=current_row1, column=3).fill = PatternFill(start_color=hex_val, end_color=hex_val, fill_type="solid")
                
                for col_index, occ_val in row_occs1.items():
                    ws1.cell(row=current_row1, column=col_index).fill = get_occ_fill1(occ_val)
                    
            # ==========================================
            # TAB 2: Occupancy Stats
            # ==========================================
            ws2 = wb.create_sheet(title="Occupancy Stats")
            
            # Write Metadata 
            ws2.append([f"Filename: {out_filename}"])
            ws2.append([f"Reference: {ref_display}"])
            ws2.append([f"Global Conserved (>{int(global_min*100)}%)"])
            ws2.append(global_list if global_list else ["None"])
            ws2.append([])
            
            # Write Headers 
            ws2.append(["Occupancy Matrix"])
            headers2 = ["Subset Name", "Count", "Hex Color", "Min Len", "Max Len", "Avg Len", "Std Dev", ""] + [f"#{c}" for c in all_occ_labels]
            ws2.append(headers2)
            
            try:
                occ_cmap2 = cm.get_cmap('Greens')
            except AttributeError:
                occ_cmap2 = matplotlib.colormaps['Greens']
                
            def get_occ_fill2(occ_value):
                rgba = occ_cmap2(occ_value)
                hex_color = mcolors.to_hex(rgba)[1:].upper()
                return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

            # Write Global Row 
            global_freq_row2 = [
                "Global Stats", 
                total_global_seqs, 
                "-",
                g_min, 
                g_max, 
                round(g_avg, 1), 
                round(g_std, 1),
                ""
            ]
            
            g_occ_dict2 = {}
            for col in all_occ_labels:
                global_freq_row2.append("") # Keep text blank
                if col in g_stats:
                    g_occ_dict2[col] = g_stats[col][2] # Occupancy is the 3rd item in the tuple
                else: 
                    g_occ_dict2[col] = 0.0
                    
            ws2.append(global_freq_row2)
            g_row_idx2 = ws2.max_row
            
            for c_idx, col in enumerate(all_occ_labels):
                col_letter_idx = c_idx + 9 
                ws2.cell(row=g_row_idx2, column=col_letter_idx).fill = get_occ_fill2(g_occ_dict2[col])
                
            ws2.append([]) # Blank row below Global Stats

            # Write Subset Rows
            last_type = None
            for res in cluster_results:
                if last_type == 'cluster' and res['type'] == 'group':
                    ws2.append([]) # Blank row separating Clusters and Groups
                last_type = res['type']
                row2 = [
                    res['name'], 
                    res['count'], 
                    res['hex'],
                    res['min'],
                    res['max'],
                    round(res['avg'], 1),
                    round(res['std'], 1),
                    ""
                ]
                
                row_occs2 = {}
                for c_idx, col in enumerate(all_occ_labels): 
                    row2.append("") # Keep text blank
                    row_occs2[c_idx + 9] = res['occ_data'].get(col, 0.0) 
                        
                ws2.append(row2)
                current_row2 = ws2.max_row
                
                if res['hex'] != "-":
                    hex_val = res['hex'].replace("#", "").upper()
                    ws2.cell(row=current_row2, column=3).fill = PatternFill(start_color=hex_val, end_color=hex_val, fill_type="solid")
                
                for col_index, occ_val in row_occs2.items():
                    ws2.cell(row=current_row2, column=col_index).fill = get_occ_fill2(occ_val)

            wb.save(out_path)
            
            msg = f"Exported to {out_path}"
            viewer.console_text.text = msg
            print(msg)
            if os.name == 'nt':
                try: os.startfile(out_dir)
                except: pass
        except Exception as e:
            viewer.console_text.text = f"IO Error: {e}"

    except Exception as e:
        viewer.console_text.text = f"Error: {e}"
        traceback.print_exc()